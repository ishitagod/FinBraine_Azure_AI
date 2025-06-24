"""
This code sample shows Prebuilt Layout operations with the Azure AI Document Intelligence client library.
The async versions of the samples require Python 3.8 or later.

To learn more, please visit the documentation - Quickstart: Document Intelligence (formerly Form Recognizer) SDKs
https://learn.microsoft.com/azure/ai-services/document-intelligence/quickstarts/get-started-sdks-rest-api?pivots=programming-language-python
"""
from azure.core.credentials import AzureKeyCredential
from azure.ai.documentintelligence import DocumentIntelligenceClient
from azure.ai.documentintelligence.models import AnalyzeDocumentRequest
from PyPDF2 import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import os
import json

AZURE_ENDPOINT = "https://formrecogocr.cognitiveservices.azure.com/"
AZURE_KEY      = "CaclGvzIwuOdWBnDsm5BSuQmZfNJug3xh1O1BJuul08jb8aYwH6nJQQJ99BFACGhslBXJ3w3AAALACOG6Egv"
PDF_FILE_PATH  = r"D:\bitsg\Finbraine_Intern\Azure AI\PDFs\PRAKASH MAKKAN CHAUDHARI.pdf"

# -----------------------------------------------------------------------------  
# 1. Count pages and build pages string
# -----------------------------------------------------------------------------
reader        = PdfReader(PDF_FILE_PATH)
num_pages     = len(reader.pages)
pages_string  = f"1-{num_pages}"   # e.g. "1-10"  

print(f"PDF has {num_pages} pages; requesting pages='{pages_string}'")

# -----------------------------------------------------------------------------  
# 2. Submit single analyze request for all pages
# -----------------------------------------------------------------------------
with open(PDF_FILE_PATH, "rb") as f:
    pdf_bytes = f.read()

client = DocumentIntelligenceClient(
    endpoint=AZURE_ENDPOINT,
    credential=AzureKeyCredential(AZURE_KEY)
)

poller = client.begin_analyze_document(
    model_id="prebuilt-layout",
    body=AnalyzeDocumentRequest(bytes_source=pdf_bytes),
    pages=pages_string
)
result = poller.result()

# -----------------------------------------------------------------------------  
# 3. If service returns fewer pages than expected, fall back to per-page calls
# -----------------------------------------------------------------------------
if len(result.pages) < num_pages:
    print(f"Warning: got {len(result.pages)} pages back; falling back to per-page requests.")
    all_pages  = []
    all_tables = []
    for p in range(1, num_pages + 1):
        sub_poller = client.begin_analyze_document(
            model_id="prebuilt-layout",
            body=AnalyzeDocumentRequest(bytes_source=pdf_bytes),
            pages=str(p)
        )
        resp = sub_poller.result()
        all_pages.extend(resp.pages)
        all_tables.extend(resp.tables)
    result.pages  = all_pages
    result.tables = all_tables

# -----------------------------------------------------------------------------  
# 4. Extract lines and tables
# -----------------------------------------------------------------------------
all_lines  = []
all_tables = []

for page in result.pages:
    for idx, line in enumerate(page.lines):
        all_lines.append({
            "page_number":  page.page_number,
            "line_number":  idx + 1,
            "content":      line.content,
            "bounding_box":[str(pt) for pt in getattr(line, "polygon", [])]
        })

for ti, table in enumerate(result.tables):
    br       = table.bounding_regions[0] if table.bounding_regions else None
    page_num = br.page_number if br else 1
    cells    = []
    for cell in table.cells:
        cells.append({
            "row_index":    cell.row_index,
            "column_index": cell.column_index,
            "content":      cell.content,
            "is_header":    getattr(cell, "kind", "") == "columnHeader"
        })
    all_tables.append({
        "table_number": ti + 1,
        "page_number":  page_num,
        "row_count":    table.row_count,
        "column_count": table.column_count,
        "cells":        cells
    })

print(f"Extracted {len(result.pages)} pages, {len(all_lines)} lines, {len(all_tables)} tables.")

# -----------------------------------------------------------------------------  
# 5. Save helpers (Excel, JSON, Markdown)
# -----------------------------------------------------------------------------
def save_to_excel(data, filename="document_analysis.xlsx"):
    wb = Workbook()
    # remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # tables → separate sheets
    for ti, table in enumerate(data["tables"]):
        ws = wb.create_sheet(f"Table_{ti+1}")
        # header row
        headers = {c["column_index"]: c["content"] for c in table["cells"] if c["row_index"] == 0}
        if headers:
            ws.append([headers.get(i, f"Column_{i}") for i in range(table["column_count"])])
        # data rows
        for r in range(1, table["row_count"]):
            row = [next((c["content"] for c in table["cells"]
                         if c["row_index"] == r and c["column_index"] == col), "")
                   for col in range(table["column_count"])]
            if any(row):
                ws.append(row)

    # pages → one sheet
    ws = wb.create_sheet("Pages")
    ws.append(["Page Number", "Width", "Height", "Unit", "Angle"])
    for page in data["pages"]:
        ws.append([
            page["page_number"],
            page["width"],
            page["height"],
            page["unit"],
            page["angle"]
        ])

    # formatting
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max((len(str(cell.value)) for cell in col), default=0)
            sheet.column_dimensions[col[0].column_letter].width = min((max_len + 2) * 1.2, 50)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    wb.save(filename)
    print(f"Excel saved to {filename}")

def save_table_to_json(data, filename="tables_data.json"):
    formatted = []
    for table in data["tables"]:
        # reorganize per row
        rows = {}
        for c in table["cells"]:
            rows.setdefault(c["row_index"], {})[c["column_index"]] = c["content"]
        header = rows.get(0, {})
        for r, cols in rows.items():
            if r == 0: continue
            entry = {header.get(0, "Heading"): cols.get(0, "")}
            for ci, val in cols.items():
                if ci == 0: continue
                entry[header.get(ci, f"Column_{ci}")] = val
            formatted.append(entry)
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(formatted, f, ensure_ascii=False, indent=2)
    print(f"JSON saved to {filename}")



# build result dict
result_dict = {
    "pages": [
        {
            "page_number": page.page_number,
            "angle":       getattr(page, "angle", 0.0),
            "width":       getattr(page, "width", 0.0),
            "height":      getattr(page, "height", 0.0),
            "unit":        getattr(page, "unit", "pixel"),
            "lines":       [{"content": l.content,
                             "bounding_box": [str(pt) for pt in getattr(l, "polygon", [])]}
                            for l in page.lines]
        }
        for page in result.pages
    ],
    "all_lines":    all_lines,
    "tables":       all_tables,
    "total_tables": len(all_tables),
    "total_lines":  len(all_lines)
}

# save outputs
save_to_excel(result_dict)
save_table_to_json(result_dict)